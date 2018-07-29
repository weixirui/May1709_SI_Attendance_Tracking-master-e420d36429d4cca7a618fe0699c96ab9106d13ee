#todo: fix all styling
#get rid of the long string of numbers
#output logs at /storage/app
#python output is in ~/Laravel/public/tmp.txt

import sys
import time
import os


# This path is needed for python modules on the SERVER #
sys.path.append('/opt/rh/python27/root/usr/lib/python2.7/site-packages/')

from openpyxl import *
from openpyxl.styles import * # why?		I have no idea, but the program won't work without this line
from boxsdk import Client, OAuth2


CURRENTTIME=time.time()
DEVMODE=False

# Command Line Args
names=sys.argv[1]
username=sys.argv[2]
duration= int(sys.argv[3])
ACCESS_TOKEN = sys.argv[4]
try:
	dm=sys.argv[5]
	if dm=="DEV":
		DEVMODE=True
except:
	DEVMODE=False

# Excel Document Style Constants #
thin_border = Border( left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin') )
thick_border = Border( left=Side(style=None), right=Side(style=None), top=Side(style='medium'), bottom=Side(style='medium') )
vert_border = Border( left=Side(style='medium'), right=Side(style='medium'), top=Side(style='thin'), bottom=Side(style='thin') )
yellowFill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
redFill = PatternFill(start_color='D99694', end_color='D99694', fill_type='solid')
whiteFill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
grayFill = PatternFill(start_color='BFBFBF', end_color='BFBFBF', fill_type='solid')
brownFill = PatternFill(start_color='DDD9C3', end_color='DDD9C3', fill_type='solid')
bold=Font(bold=True)


# Login Credentials #
CLIENT_ID = "ka45qzs9e9cnabzmaaxmmi2mrpll9j32"#"akbuknoxh68mlusnhqio5n4lh43emdxf"
CLIENT_SECRET = "T8Oxv8xK4ObJY6SqMzLQXZIucFxGqlvs"#"WCxlaGEehP9itIwzLV1oJ2CYrsm9Kb4l"
oauth2 = OAuth2(CLIENT_ID, CLIENT_SECRET, access_token=ACCESS_TOKEN)
client = Client(oauth2)
workingFolder='0'

def download_file():
	fid = '0'

	# All of the items in the box user's 'root' directory.
	items = client.folder(folder_id=fid).get_items(limit=100, offset=0)


	# Find the appropriate directory
	for x in items:
		if "Training" in str(x):
			fid = int(str(x).split("(")[0].split()[-1])
	fid=22073241963
	workingFolder = fid
	items = client.folder(folder_id=fid).get_items(limit=100, offset=0)


	# Find Excel file for user
	for x in items:
		if username in str(x):
			fid=int(str(x).split("(")[0].split()[-1])
			classAndSID=str(x).split("-")[-1][0:-2]
			fileName="Attendance Tracking Spreadsheet -"+classAndSID


	# Open the intended file
	with open("../public/input-"+fileName, 'wb') as open_file:
	    client.file(fid).download_to(open_file)
	fileNumber=fid
	return fileName,workingFolder,fileNumber

# Upload the document back to box.
def uploadFile(fileName,workingFolder,fid):
	root_folder = client.folder(folder_id=workingFolder)
	file_path = os.path.join(os.path.dirname(os.path.realpath(__file__)), fileName)

	client.file(fid).update_contents(fileName)

# read the names from the file given by the command line arguments
def getPeople():
	file = open(names,"r")
	namesToProcess=file.readlines()
	people=[]
	for x in namesToProcess:
		people=people+[[x.split()[0].upper(),x.split()[1].upper(),x.split()[-1]]]
	return people

# read the excel file and produce a list of names from the cells of column B
def scanExcelForNames(ws):
	n=[]
	i=6
	while i!=-1:
		i=i+1
		item=ws["B"+str(i)].internal_value
		n=n+[item]
		if item==None:
			for x in range(len(n)-1):
				n[x]=str(n[x]).split()
			del n[-1]
			return n

# find the amount of lines that the table has to be extended by
def getAmountToExtend(names1,names2):
	namesToAdd=[]
	i=0
	for x in names1:
		j=i;
		for y in names2:
			if x[0]==y[0] and x[1]==y[1]:
				i=i+1
		if j==i:
			namesToAdd=namesToAdd+[x]
	return len(namesToAdd),namesToAdd

# add the names from the list given by the command line arguments to column B
def addNames(ws,namesToAdd,startAt):
	for i in range(len(namesToAdd)):
		ws["B"+str(i+startAt)]=namesToAdd[i][0]+' '+namesToAdd[i][1]
		ws["B"+str(i+startAt)].alignment = Alignment(horizontal='center')
		ws["B"+str(i+startAt)].fill = yellowFill
		ws["B"+str(i+startAt)].border=thin_border

# extend column A
def extendFirstColumn(ws,amountToExtend,startAt):
#	print startAt,amountToExtend
	s = ws["A8"].internal_value
	for i in range(startAt,amountToExtend+startAt):
		ws["A"+str(i)]=s.replace("8",str(i))
		ws["A"+str(i+1)].fill=redFill
		ws["A"+str(i)].fill=whiteFill
		ws["A"+str(i)].alignment = Alignment(horizontal='center')
		ws["A"+str(i)].border=thin_border
		ws["A"+str(i+1)].border=thin_border
#		ws["A"+str(i)].font = cell.font.copy(bold=True)

# Returns the letter of the column given as num. EX: getletter(1)="A", getletter(27)="AA"
def getLetter(num):
	num = num-1
	out = chr(ord('A')+num/26-1)+chr(ord('A')+num%26)
	if out[0] == '@':
		out=out[1]
	return out

# Get document column number for the given date.
def getNumFromDate(date):
	if int(date.split('-')[1])==4:
		return 82+int(date.split('-')[2])
	if int(date.split('-')[1])==3:
		return 50+int(date.split('-')[2])
	if int(date.split('-')[1])==2:
		return 21+int(date.split('-')[2])
	# return 21+int(date.split('-')[2])


def extendBorders(ws,amountToExtend,startAt):
	key="11111111112111111211131112111111211111121111112111311121111112222222211111121111113211111121111112111111211111124444440000000"
	for j in range(1,119):
		for i in range(startAt,amountToExtend+startAt):
			ws[getLetter(j)+str(i)].border=thin_border
			ws[getLetter(j)+str(i)].alignment = Alignment(horizontal='center')
			if key[j]=='1':
				pass
			if key[j]=='2':
				ws[getLetter(j)+str(i)].fill=grayFill
			if key[j]=='3':
				for k in {0,1}:
					ws[getLetter(j)+str(i-k)]="=SUM("+getLetter(j)+"7:"+getLetter(j)+str(i-1-k)+")"#THIS IS WRONG
					ws[getLetter(j)+str(i-k)].border=vert_border
					ws[getLetter(j)+str(i-k)].fill=whiteFill
			if key[j]=='4':
				ws[getLetter(j)+str(i)].fill=brownFill
	for j in range(6,119):
		ws[getLetter(j)+str(amountToExtend+startAt)].border=thick_border
		ws[getLetter(j)+str(amountToExtend+startAt)]="=SUM("+getLetter(j)+"7:"+getLetter(j)+str(amountToExtend+startAt-1)+")"
		ws[getLetter(j)+str(startAt)]=""

# add in the correct number of hours for each cell
def addHours(ws,people,length):
	for i in range(length):
		for j in range(len(people)):
			if(ws["B"+str(i+7)].internal_value.split()[0]==people[j][0]):
				if(ws["B"+str(i+7)].internal_value.split()[1]==people[j][1]):
					ws[getLetter(getNumFromDate(people[j][-1]))+str(i+7)]=duration

def replaceChart(ws,numNames,extendAmount):
	#set i to the index of the top left corner of the statistics chart
#	print str(numNames+6)
	index=1
	while ws[getLetter(index)+"1"].internal_value!="Institution:" and index<200:
		index=index+1
	for i in range(1,20):
		for j in range(1,40):
			ws[getLetter(i+index)+str(j)]=1
#			try:
#				ws[getLetter(i+index)+str(j)]=ws[getLetter(i+index)+str(j)].internal_value.replace(str(numNames+6),str(numNames+6)+extendAmount+1000)
#				ws[getLetter(i+index)+str(j)]=ws[getLetter(i+index)+str(j)].internal_value.replace(str(numNames+6)+1,str(numNames+6)+extendAmount+1001)
#			except:
#				pass

print "Start"

fileName="input_test.xlsx"
if DEVMODE==False:
	fileName,workingFolder,fid=download_file()

print "File Downloaded "+str(time.time()-CURRENTTIME)
CURRENTTIME=time.time()

wb = load_workbook('../public/'+"input-"+fileName)
ws = wb.active


print "Workbook Loaded "+str(time.time()-CURRENTTIME)
CURRENTTIME=time.time()

people=getPeople()
excelNames=scanExcelForNames(ws)

print "People Loaded "+str(time.time()-CURRENTTIME)
CURRENTTIME=time.time()

extendAmount,namesToAdd=getAmountToExtend(people,excelNames)
addNames(ws,namesToAdd,len(excelNames)+7)
extendFirstColumn(ws,extendAmount,len(excelNames)+7)
extendBorders(ws,extendAmount,len(excelNames)+7)
addHours(ws,people,extendAmount+len(excelNames))
replaceChart(ws,len(excelNames),extendAmount)


print "Worksheet Processed "+str(time.time()-CURRENTTIME)
CURRENTTIME=time.time()

#print people
#print excelNames
#print str(len(excelNames))+" "+str(extendAmount)
#print namesToAdd

#print ws["J11"].font.italic



	
file = open(fileName, 'a+')
file.close()
if DEVMODE==True:

	wb.save("out"+fileName)

if DEVMODE == False:
	wb.save(fileName)

print "File Saved "+str(time.time()-CURRENTTIME)
CURRENTTIME=time.time()

if DEVMODE == False:
	uploadFile(fileName,workingFolder,fid)
	os.remove(fileName)


print "File Uploaded "+str(time.time()-CURRENTTIME)
CURRENTTIME=time.time()



