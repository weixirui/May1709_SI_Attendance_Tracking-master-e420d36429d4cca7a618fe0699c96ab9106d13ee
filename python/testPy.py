#TODO: Fix getNumFromDate

#nameList	1
#username	2
#duration	3
#refresh key	4

from openpyxl import *
from openpyxl.styles import * # why?		I have no idea, but the program won't work without this line
from boxsdk import Client, OAuth2
import time
import sys
import os

# Predefine Arrays #
namelist = []
datelist = []
namesused = []
namesnotused = []

# Command line args assigned to variables
names=sys.argv[1]
username=sys.argv[2]
duration=sys.argv[3]
ACCESS_TOKEN = sys.argv[4]


# Excel Document Style Constants #
thin_border = Border( left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin') )
thick_border = Border( left=Side(style=None), right=Side(style=None), top=Side(style='medium'), bottom=Side(style='medium') )
yellowFill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
grayFill = PatternFill(start_color='BFBFBF', end_color='BFBFBF', fill_type='solid')


# Login Credentials #
CLIENT_ID = "akbuknoxh68mlusnhqio5n4lh43emdxf"
CLIENT_SECRET = "WCxlaGEehP9itIwzLV1oJ2CYrsm9Kb4l"
oauth2 = OAuth2(CLIENT_ID, CLIENT_SECRET, access_token=ACCESS_TOKEN)
client = Client(oauth2)
workingFolder='0'

def download_file():
	fid = '0'

	# All of the items in the box user's 'root' directory.
	items = client.folder(folder_id=fid).get_items(limit=100, offset=0)


	# Find the appropriate directory
	for x in items:
		if "Senior Design Test" in str(x):
			fid = int(str(x).split("(")[0].split()[-1])
	workingFolder = fid
	items = client.folder(folder_id=fid).get_items(limit=100, offset=0)


	# Find Excel file for user
	for x in items:
		if username in str(x):
			fid=int(str(x).split("(")[0].split()[-1])
			classAndSID=str(x).split("-")[-1][0:-2]
			fileName="Attendance Tracking Spreadsheet -"+classAndSID


	# Open the intended file
	with open("input-"+fileName, 'wb') as open_file:
	    client.file(fid).download_to(open_file)
	return fileName,workingFolder


# Mark document with hours for all students present
def markhours():
	i = 7
	while ws[i][1].internal_value != None:
		name = ws[i][1].internal_value.split()[0:2]
		name = name[0]+' '+name[1]
		for x in range(len(namelist)):
			if name == namelist[x]:
				namesused.append(name)
				ws[getLetter(getNumFromDate(datelist[x]))+str(i)] = duration
				#ws[getLetter(getNumFromDate(datelist[x]))+str(i)].fill=yellowFill
		i=i+1
	return i


# Add missing names to the excel document
def addNewNames(i):
	namesnotused = namelist
	s = ws["A8"].internal_value
	for x in namesused:
		namesnotused.remove(x)
	for x in range(6,118):
		ws[getLetter(x)+str(i+len(namesnotused))] = "=SUM("+getLetter(x)+"7:"+getLetter(x)+str(i+len(namelist)-1)+")"
		ws[getLetter(x)+str(i+len(namesnotused))].border = thick_border

	for x in range(len(namesnotused)):
		ws["A"+str(i+x+1)] = "=SUM(A7:A"+str(i+x)+")"
		ws["A"+str(i+x+1)].alignment = Alignment(horizontal='center')
		ws["B"+str(i+x)] = namesnotused[x]
		ws["B"+str(i+x)].alignment = Alignment(horizontal='center')
		ws["B"+str(i+x)].fill = yellowFill
		ws["A"+str(i+x)] = s.replace("8",str(i+x))
		for y in range(10, 118):
			#print(getLetter(y)+str(x+i))
			sunList = {25, 32, 39, 46, 54, 61, 68, 75, 83, 90, 97, 104, 111, 118}
			if y in sunList:
				ws[getLetter(y)+str(x+i)].fill = grayFill
		for y in range(6,118):
			ws[getLetter(y)+str(x+i)].value = None
			#print(ws[getLetter(y)+str(x+i+1)].fill)#=ws[getLetter(y)+str(x+i)].fill
		for y in range(1,118):
			ws[getLetter(y)+str(x+i)].border = thin_border
		ws["AX"+str(i+x)] = ws["AX"+str(9)].internal_value.replace("9",str(i+x))
		#ws["AX"+str(i+x)]=ws["AX"+str(9)].alignment = Alignment(horizontal='center')
	for x in range(len(namesnotused)):
		ws[getLetter(getNumFromDate(datelist[-1-x]))+str(i+len(namelist)-1-x)] = duration
		ws[getLetter(getNumFromDate(datelist[-1-x]))+str(i+len(namelist)-1-x)].alignment = Alignment(horizontal='center')
		#ws[getLetter(getNumFromDate(datelist[-1-x]))+str(i+len(namelist)-1-x)].fill=yellowFill


# Upload the document back to box.
def uploadFile(fileName,workingFolder):
	root_folder = client.folder(folder_id=workingFolder)
	file_path = os.path.join(os.path.dirname(os.path.realpath(__file__)), fileName)
	#print file_path
	oname = str(time.time())+fileName
	try:
		a_file = root_folder.upload(file_path, file_name=oname)
	except:
		try:
			a_file = root_folder.upload(file_path, file_name=oname)
		except:
			try:
				a_file = root_folder.upload(file_path, file_name=oname)
			except:
				print "bad connection"


# Returns the letter of the column given as num. EX: getletter(1)="A", getletter(27)="AA"
def getLetter(num):
	num = num-1
	out = chr(ord('A')+num/26-1)+chr(ord('A')+num%26)
	if out[0] == '@':
		out=out[1]
	return out


# Get document column number for the given date.
def getNumFromDate(date):
	return 50+int(date.split('-')[2])
	# return 21+int(date.split('-')[2])


# Prepare arrays of names and dates found in the input document.
def makelists():
	file = open(names,"r")
	inputFile = file.readlines()
	for x in range(len(inputFile)):
		inputFile[x] = inputFile[x].upper()

	for i in range(len(inputFile)):
		namelist.append(inputFile[i].split()[0]+' '+inputFile[i].split()[1])
		datelist.append(inputFile[i].split()[-1])



##########################################
## This is where stuff actually happens ##
##########################################

#'''
#fileName,workingFolder=download_file()
#
##wb = load_workbook("input-"+fileName)
#wb = load_workbook('../Python/'+"input-"+fileName)
#ws = wb.active##

#makelists()
#addNewNames(markhours())
#
##file = open(fileName, 'a+')
#file = open('../Python/'+fileName, 'a+')
#file.close()
#wb.save(fileName)
#uploadFile(fileName,workingFolder)
#os.remove('../Python/'+"input-"+fileName)
#os.remove('../Python/'+fileName)
#'''and None
open_file=open(names,"r")
client.file(144656917590).update_contents(names)

##########################################
##        End of stuff happening        ##
##########################################








